from tkinter import *
import translators as ts
from tkinter import messagebox
from PIL import ImageTk, Image
import pyttsx3
from  deep_translator import GoogleTranslator
from openpyxl.workbook import Workbook
from openpyxl import load_workbook



def strt() :
    root.withdraw()
    win = Tk()
    win.title("TRANSLATOR")
    win.geometry('600x500')
    win.iconbitmap('language.ico')
    win.configure(bg='black')




    def back() :
        root.deiconify()
        win.destroy()



    def text_to_speech(val):
        t = pyttsx3.init()
        t.setProperty("rate", 140)
        t.say(val)
        t.runAndWait()




    def French():
      word = T.get("1.0","end-1c")
      #translation =ts.google(word,from_language = 'en', to_language='fr')
      translation = GoogleTranslator(source='auto', target = 'fr',).translate(word)
      l2.config(text=f'TRANSLATED WORD  : {translation}')
      text_to_speech(translation)

    def german():
      word = T.get("1.0","end-1c")
      #translation =ts.google(word,from_language = 'en', to_language='de')
      translation = GoogleTranslator(source='auto', target='de', ).translate(word)
      l2.config(text=f'TRANSLATED WORD  : {translation}')
      text_to_speech(translation)


    def spanish():
      word = T.get("1.0","end-1c")
      #translation = ts.google(word, from_language='en', to_language='es')
      translation = GoogleTranslator(source='auto', target='es', ).translate(word)
      l2.config(text=f'TRANSLATED WORD  : {translation}')
      text_to_speech(translation)


    def hindi():
      word = T.get("1.0", "end-1c")
      #translation =ts.google(word,from_language = 'en', to_language='hi')
      translation = GoogleTranslator(source='auto', target='hi', ).translate(word)
      l2.config(text=f'TRANSLATED WORD  : {translation}')
      text_to_speech(translation)


    def turkish():
      word = T.get("1.0","end-1c")
      #translation =ts.google(word,from_language = 'en', to_language='tr')
      translation = GoogleTranslator(source='auto', target='tr', ).translate(word)
      l2.config(text=f'TRANSLATED WORD  : {translation}')
      text_to_speech(translation)


    def chinese():
      word = T.get("1.0","end-1c")
      #translation = ts.google(word, from_language='en', to_language='zh')
      translation = GoogleTranslator(source='auto', target='zh', ).translate(word)
      text_to_speech(translation)


    def Arabic():
      word = T.get("1.0","end-1c")
      #translation =ts.google(word,from_language = 'en', to_language='ar')
      translation = GoogleTranslator(source='auto', target = 'ar',).translate(word)
      text_to_speech(translation)


    def Persian():

      word = T.get("1.0","end-1c")
      #translation =ts.google(word,from_language = 'en', to_language='fa')
      translation = GoogleTranslator(source='auto', target='fa', ).translate(word)
      l2.config(text=f'TRANSLATED WORD  : {translation}')
      text_to_speech(translation)






    l1=Label(win,text="Enter your text Here!!",font=('stencil',15,'italic'),fg='green',bg="black")
    l1.place(x=40,y=15)
    l3=Label(win,text="Translation:",font=('stencil',15,'italic'),fg='green',bg="black")
    l3.place(x=5,y=325)
    T=Text(win,height=17, width=37, bg="yellow")
    T.place(x=20,y=40)



    '''
    Btn1=Button(win,text='French',bd=10,padx=8,pady=8,bg="powder blue",width=6,command=French)
    Btn1.place(x=5,y=92)
    Btn2=Button(win,text='German',bd=10,padx=8,pady=8,bg="yellow",width=6,command=german)
    Btn2.place(x=85,y=92)
    Btn3=Button(win,text='Spanish',bd=10,padx=8,pady=8,bg="blue",width=6,command=spanish)
    Btn3.place(x=165,y=92)
    Btn4=Button(win,text='Hindi',bd=10,padx=8,pady=8,bg="Green",width=6,command=hindi)
    Btn4.place(x=240,y=92)
    Btn5=Button(win,text='Turkish',bd=10,padx=8,pady=8,bg="Pink",width=6,command=turkish)
    Btn5.place(x=5,y=152)
    Btn6=Button(win,text='Chinese',bd=10,padx=8,pady=8,bg="orange",width=6,command=chinese)
    Btn6.place(x=85,y=152)
    Btn7=Button(win,text='Arabic',bd=10,padx=8,pady=8,bg="white",width=6,command=Arabic)
    Btn7.place(x=165,y=152)
    Btn8=Button(win,text='Persian',bd=10,padx=8,pady=8,bg="DodgerBlue2",width=6,command=Persian)
    Btn8.place(x=245,y=152)'''



    l2 = Label(win, text="", font=('arial', 10, 'italic'), fg="cyan", bg="black")
    l2.place(x=15, y=350)


    langs = [ "French", "Spanish", "Hindi", "Chinese", "German", "Turkish", "Arabic", "Persian"]
    def cnvrt(event):
        st = clicked.get()
        if st == "French" :
            French()

        elif st == "German" :
            german()

        elif st == "Spanish":
            spanish()

        elif st == "Hindi":
            hindi()

        elif st == "Chinese":
            chinese()

        elif st == "Turkish":
            turkish()


        elif st == "Arabic":
            Arabic()

        elif st == "Persian":
            Persian()

        else :
            l2.config(text = "Appropriate Language not selected : ")








    clicked = StringVar()
    clicked.set("Choose any one")
    drop = OptionMenu(win, clicked, *langs,  command = cnvrt)
    drop.config(bg="green", fg="white")
    drop["menu"].config(bg="white", fg = "green")
    drop.place(x=370, y=5)






    def on_enterbtn9(e):
        btn9['bg'] = 'violet'

    def on_leavebtn9(e):
        btn9['bg'] = 'white'


    btn9=Button(win, text="BACK", fg="white", bg="green",width=7, height=2,bd=15, command=back)
    btn9.place(x=400, y=350)
    btn9.bind('<Enter>', on_enterbtn9)
    btn9.bind('<Leave>', on_leavebtn9)






root = Tk()
root.geometry('600x500')
root.title('Translater')
root.iconbitmap('translator.ico')
root.configure(bg='white')


def login():
    new = Toplevel(root)
    new.geometry("400x400")
    new.title("Login")
    new.iconbitmap('login.ico')
    new.configure(bg='white')

    C1 = Canvas(new, bg="gold2", height=60, width=400)
    C1.place(x=0, y=0)
    canlab = Label(new, text = "Login To Your Account", font=('Agency FB', 25, 'italic bold'), bg="gold2", fg="red")
    canlab.place(x=15, y=15)
    C2 = Canvas(new, bg="gold2", height=100, width=400)
    C2.place(x=0, y=340)



    lab1 = Label(new, text="Name" , bg='white', fg="red", font=('Arial Rounded MT Bold', 15, 'italic  '))
    lab1.place(x=100, y=80)
    E1 = Text(new, height=2, width=25, bg="gold2", fg='red', font=('Ariel', 12, 'italic  ') )
    E1.place(x=100, y=110)

    lab2 = Label(new, text="Email",bg='white',  fg="red", font=('Arial ', 15, 'italic  '))
    lab2.place(x=100, y=160)
    E2 = Text(new,height=2, width=25,  bg="gold2", fg='red',font=('Ariel', 12, 'italic  '))
    E2.place(x=100, y=190)

    wb = load_workbook('record.xlsx')
    ws = wb.active
    col_b=ws['B']
    col_c=ws['C']





    def submit():
        data1 = E1.get("1.0", "end-1c")
        data2 = E2.get("1.0", "end-1c")
        global flag1,flag2
        flag1=0
        flag2=0
        for item in col_c :

            if data1 == item.value:

                flag1=1
                break

        for item in col_b :

            if data2 == item.value:

                flag2=1
                break


        def rtrn():

            lab_cong.config(text="Congrats !!  " + data1 + "  Your Account successfully logged in")

            def on_enterbtn_start(e):
                btn_start['fg'] = 'white'
                btn_start['bg'] = 'red'

            def on_leavebtn_start(e):
                btn_start['fg'] = 'white'
                btn_start['bg'] = 'gold2'


            btn_start = Button(root, text="START", fg="white", bg="gold2", activebackground='green', width=7, height=2,
                               font=('stencil', 10, 'italic  '), bd=5, command=strt)
            btn_start.place(x=500, y=260)
            btn_start.bind('<Enter>', on_enterbtn_start)
            btn_start.bind('<Leave>', on_leavebtn_start)

            new.withdraw()






        if flag1 ==1 and flag2 ==1:
            rtrn()

        else :
            lab_cong.config(text="Either your entered details are wrong or you account not exist :")




    def on_enterb(e):
        btn_b['bg'] = 'red'
        btn_b['fg'] = 'white'

    def on_leaveb(e):
        btn_b['bg'] = 'gold2'
        btn_b['fg'] = 'red'

    btn_b = Button(new, text="SUBMIT", fg="red", bg="gold2", activebackground='green', width=8, height=2,
                       font=('stencil', 12, 'italic   '), bd=5, command=submit)
    btn_b.place(x=230, y=250)
    btn_b.bind('<Enter>', on_enterb)
    btn_b.bind('<Leave>', on_leaveb)


















C = Canvas(root, bg="black",height=300,width=350)
C.place(x=50,y=100)
global our_images, count
count=0

our_images = [
               ImageTk.PhotoImage((Image.open("g1.png")).resize((350,300))),
               ImageTk.PhotoImage((Image.open("q.png")).resize((350,300))),
               ImageTk.PhotoImage((Image.open("s.png")).resize((350,300))),
               ImageTk.PhotoImage((Image.open("o.png")).resize((350, 300))),
               ImageTk.PhotoImage((Image.open("d.png")).resize((350, 300))),
               ImageTk.PhotoImage((Image.open("m.png")).resize((350, 300)))
    ]

C.create_image(0, 0, anchor=NW, image=our_images[0])



def next():
    global count
    if count == 5 :
        C.create_image(0, 0, anchor=NW, image=our_images[0])
        count = 0

    else :
        C.create_image(0, 0, anchor=NW, image=our_images[count+1])
        count+=1

    root.after(1000, next)

next()

global title
title=Label(root,text=''' MULTIPLE \n LANGUAGE TRANSLATOR ''',font=('stencil', 24, 'italic bold  '),fg='gold2',bg="white")
title.place(x=70,y=25)


def on_enterlogin_start(e):
    btn_login['fg'] = 'white'
    btn_login['bg'] = 'red'


def on_leavelogin_start(e):
    btn_login['fg'] = 'white'
    btn_login['bg'] = 'gold2'

btn_login = Button(root, text="LOGIN", fg="white",bg="gold2", activebackground='green',width=7, height=2,font=('stencil', 10, 'italic  ') , bd=5, command=login)
btn_login.place(x=500, y=50)
btn_login.bind('<Enter>', on_enterlogin_start)
btn_login.bind('<Leave>', on_leavelogin_start)


def main_exit():
    rr = messagebox.askyesnocancel('EXIT', 'Are you want to exit !', parent=root)
    if (rr == True):
     root.destroy()



def on_enterexit_btn(e):
    exit_btn['bg'] = 'red'
    exit_btn['fg'] = 'white'


def on_leaveexit_btn(e):
    exit_btn['bg'] = 'gold2'
    exit_btn['fg'] = 'white'



exit_btn = Button(root, text='Exit', bd=5, fg="white",bg="gold2", activebackground='green', width=7, height=2,font=('stencil',10 , 'italic  '), compound=BOTTOM,command=main_exit)
exit_btn.place(x=500, y=360)
exit_btn.bind('<Enter>', on_enterexit_btn)
exit_btn.bind('<Leave>', on_leaveexit_btn)


lab_cong=Label(root,text="",font=('arial', 10, 'italic bold  '),fg='red',bg="white")
lab_cong.place(x=30,y=450)

root.mainloop()